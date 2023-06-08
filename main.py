import openpyxl
import json
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import datetime

filename = ""


# ExcelファイルをJsonにして保存する関数
def convert_to_json():
    wb = openpyxl.load_workbook(filename)

    if isEnglish:
        output_folder = filedialog.askdirectory(title="Choice output folder")

    else:
        output_folder = filedialog.askdirectory(title="出力フォルダを選択")

    if not output_folder:
        return

    name_pattern = name_entry.get().strip()

    checkbox = {ws: var.get() for ws, var in checkbox_values.items()}

    for ws in wb.worksheets:
        if not checkbox.get(ws.title):
            continue

        for table in ws.tables.values():
            table_range = table.ref

            data = []

            header = None

            for row in ws[table_range]:
                if header is None:
                    header = [cell.value for cell in row]

                else:
                    values = []

                    for cell in row:
                        value = cell.value if cell.value is not None else "null"

                        values.append(value)

                    data_row = dict(zip(header, values))

                    if "description" in data_row:
                        del data_row["description"]

                    for cell in data_row:
                        try:
                            data_row[cell] = json.loads(data_row[cell])

                        except:
                            pass

                    data.append(data_row)

            json_string = json.dumps(data, ensure_ascii=False)

            # templateで保存ファイル名を指定する
            table_name = table.name

            if name_pattern:
                output_filename = name_pattern.replace("{table_name}", table_name)

                output_filename = output_filename.replace("{sheet_name}", ws.title)

                output_filename = output_filename.replace(
                    "{today}", datetime.date.today().today().strftime("%Y_%m_%d")
                )

            else:
                output_filename = table_name

            titletext = "JSONファイルを保存"

            if isEnglish:
                titletext = "Save Json File"

            output_filename = filedialog.asksaveasfilename(
                initialdir=output_folder,
                initialfile=output_filename,
                defaultextension=".json",
                filetypes=[("JSONファイル", "*.json")],
                title=titletext,
            )

            if not output_filename:
                continue

            with open(output_filename, mode="w", encoding="utf-8") as f:
                f.write(json_string)

            print(json_string)


# guiの設定

root = tk.Tk()

root.title("Excel to JSON Converter")

# 画像を設定
data = """R0lGODlhlgCWAMIDAAAAAD9IzCKxTP///////////////////yH+EUNyZWF0ZWQg
d2l0aCBHSU1QACH5BAEKAAQALAAAAACWAJYAAAP+OLrc/jDKSau9OOvNu/9gKI5k
aZ5oqq5s675wLM90bd94ru987zuBYERALBqPyKRyyWw6n9CoVBAAWAOPqXbL7Xq7
1nD2Sy6bz82w2IFuu99a9ZoNr9vt8ivk+Otn8gBCY0V+hROAgYJ0hIaNDYhBig18
jpWQkViDRJWOgAqYmYubnIWen5h7RqR+pqeRqYyrPa2ukgyUsju0taGTqrk5iA+g
sKPANsLDqJoCJbHHh7sMxMwkSCPPyNLTy6LNIkohuDXbDdTe4Esf19pyFee+2RxO
HeE05UDd8cb19BtMM/Dle1WtHzt2GPy9EDjQ1oJxGewpQFghSotkF+Dd+qX+QeJE
ihKSfARpAmNGfRvlVRTpjV+xZx5LmDxJEF1EkilV5pQXU8RMmg5HugzZsyVRji2/
jWA4QeNDpEdxMlO6T6dQqxqYNkX5FOtOr1WpXh1qNITWrTWrWmB5ge0At0MgcjiL
NuhbqFHBhv0G9yXZP3lEOL2aUC6Fa32nis1Kl8Lgu3q/dmyy1jCGnx4eW867+LC6
lXgBByahOfSExJwjS72A+UPpyHtvbo7doTGG1389zy6r27QF27e5Ep632mZv2A+A
B0+70wPqQZV3R3OnAnfnwtInZ4+gPIN1EMX/hZcAoLz58i2+g98O2jeF8+ddqF83
PjryR/DRyxcOObf+bPf2+TddfDDMtx57U9WWnwwGHnjfV9e9tyCD/CF4Wn2KzTUh
hcx1JWBbGCalIXw1NEichZJtkB8ANpioHYAXwujAiiy2WKGMqTn4IQM04uAiiChm
WEGPPt74oIdH+hXhAivu8ONxO2LnXpNOGhklbc4kRmWVHQ4npWpX9ocUkTw8qWSA
XzJCZplWLilie2GOJcCabHYpJpA4FhUjEXTWadd2z+31YBF9+tlLc3C6iWSgENBY
ox9mQgilno1uWUikksZFCaNMWnppm2le5xanhfaBKZaLLhZoqaaC+p9LqGnJaqt2
BolQoHA5KsupCXLKlq67uvriZ5MOAGywteL+qJiyYs45qyG8KhknZMcCE+2Z/1Rr
rbAGTdupp9smm2Q8CoIb7p/MprCinatci8K6/JHibknwssvJvCT0+Ji83MZA5L73
9rvQhgMAXAm+IHhqsCMIewDuwo00PGJ5HUIMrcAqwKeRxZ+K662K+W0cb8AeK2pW
vbZwDCnGMjUpsr0Hs7zUhi/bxW/JGRNc86G5SDygeeagpDKt6I77s37cMDd0gQ7V
5HOlJDZ06NL7pUXN0/gRnHTKI0tNk3f6XC0zyFpvPfW+oKTNszIwm92L2Dj7ZO4C
Tj2mdtqO3b2226HAXfTHM85NN1d2622v4fGqPThBWAtuttds6w0B4nf+R+70Mk8D
oPnmj9bFc+ELUy550HgXjPnYQ3KuOVCfC1d63opTTrrfmUis+uqsQ+4263ZVvjjt
vLxULudgVyz067AfjrzvwBckmuq87B387GdzDDE8kjdvHGu31+i75bpHbzP174Rd
ufZqXdb973UjPz3f0j+e/NvMM4561tCL383o8ks9/vsSwF72UBGt9bHvfN/rn/8S
aLFzGM50r+CVAQ8IQcRFAGCiE0QD9xdBAnowbkfrXAUVIbsLdi2DWNig/To4QBBy
53bUq577wOe5Fv6PfdN7YKQmSMEYxi965bOgCjV4OgS6EH/Ei+Hk2jbCG5qQhW0r
He2I8aPuiRDniJlJ4HJWqLzm6RBjVlziDC3APw04UHELxKH+UiiwMD5xjG+kmgK1
iMb32XBvEOFhHENXRjO6z4h/5GDsuKXHAD7QctZD4QL7OEL4NdJ5hZzf60oYRNHt
sY5zHGTcIklGFMLxklpcpJ0EiLfvcHKLGdyAJSU5MuMV8W9uDEEqO2DBTjpRlUaK
5QgOCQ1bokuXJ2BiL8mXBSt68pjITKYyl8lMqHGOmdCMpjSnqcjkqI6a2MymNptp
zc1t85vgDCcmsybOcpozm8NMpzrXyc52uvOd8IynPOdJz3ra8574zKc+98lPeiYA
ADs=
        """

root.tk.call("wm", "iconphoto", root._w, tk.PhotoImage(data=data))


checkbox_values = {}

sheets_frame = tk.Frame(root)

sheets_frame.pack()


wb = None


# Excelファイルを開く
def open_excel_file():
    global wb
    global filename

    global checkbox_values

    # 初期化

    wb = None

    checkbox_values.clear()

    if isEnglish:
        filename = filedialog.askopenfilename(title="Choice Excel File")

    else:
        filename = filedialog.askopenfilename(title="Excelファイルを選択")

    for widget in sheets_frame.winfo_children():
        widget.destroy()

    # Excelファイルを開く

    if filename:
        wb = openpyxl.load_workbook(filename)

        for ws in wb.sheetnames:
            var = tk.BooleanVar()

            checkbox_values[ws] = var

            checkbox = tk.Checkbutton(sheets_frame, text=ws, variable=var)

            checkbox.pack(anchor="w")

        convert_button.config(state=tk.NORMAL)


# GUIの要素の初期化

open_button = tk.Button(root, text="Excelファイルを開く", command=open_excel_file)

open_button.pack(pady=10)


name_frame = tk.Frame(root)

name_frame.pack()


name_label = tk.Label(name_frame, text="ファイルを保存:")

name_label.pack(side="left")


name_entry = tk.Entry(name_frame)

name_entry.insert(tk.END, "{table_name}")

name_entry.pack(side="left")


help_button = tk.Button(root, text="?", width=2)


isEnglish = False


# 言語切り替え
def switch_language():
    global isEnglish

    current_language = language.get()

    if current_language == "日本語":
        isEnglish = False

        language.set("日本語")

        open_button.config(text="Excelファイルを開く")

        name_label.config(text="ファイルを保存:")

        help_button.config(text="?")

        convert_button.config(text="JSONに変換")

    else:
        isEnglish = True

        language.set("English")

        open_button.config(text="Open Excel File")

        name_label.config(text="File Name template:")

        help_button.config(text="?")

        convert_button.config(text="Convert To Json")


language = tk.StringVar()

language.set("日本語")


language_radiobutton = tk.Radiobutton(
    root, text="日本語", variable=language, value="日本語", command=switch_language
)

language_radiobutton.pack(anchor="w")


language_radiobutton = tk.Radiobutton(
    root, text="English", variable=language, value="English", command=switch_language
)

language_radiobutton.pack(anchor="w")


help_button.pack(pady=5)


# helpメッセージ
def show_name_template_help():
    if isEnglish:
        messagebox.showinfo(
            "Name Template Help",
            "Set the file name template. {table_name} replaces the actual table name, {sheet_name} replaces the sheet name, {today} replaces today's year_month_day. if no template is specified, the table name itself will be used as the file name.",
        )

    else:
        messagebox.showinfo(
            "名前テンプレートヘルプ",
            "ファイル名のテンプレートを設定してください。{table_name}は実際のテーブル名に置き換え、{sheet_name}はシート名に置き換えます、{today}は今日の年_月_日を置き換えます。をテンプレートが指定されていない場合、テーブル名自体がファイル名として使用されます。",
        )


help_button.config(command=show_name_template_help)


convert_button = tk.Button(
    root, text="JSONに変換", command=convert_to_json, state=tk.DISABLED
)

convert_button.pack(pady=10)


root.mainloop()
